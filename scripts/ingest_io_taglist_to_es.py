#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ingest EPMS IO taglist Excel into Elasticsearch with embeddings from Ollama.

Prereqs:
  pip install openpyxl requests

Run:
  python ingest_io_taglist_to_es.py --xlsx "IO_Address_Tag_List.xlsx" --index epms_docs
"""

import argparse, json, math, re, time
import openpyxl
import requests

def clean(x):
    if x is None: 
        return None
    x = str(x).strip()
    x = re.sub(r"\s+", " ", x)
    return x

def is_blank(v):
    return v is None or (isinstance(v, float) and math.isnan(v)) or (isinstance(v, str) and v.strip() == "")

def get_embedding(ollama_url: str, model: str, text: str, timeout: int = 60):
    url = ollama_url.rstrip("/") + "/api/embeddings"
    payload = {"model": model, "prompt": text}
    r = requests.post(url, json=payload, timeout=timeout)
    r.raise_for_status()
    data = r.json()
    emb = data.get("embedding")
    if not isinstance(emb, list):
        raise RuntimeError(f"Unexpected embedding response: {data}")
    return emb

def ensure_index(es_url: str, index: str, dims: int):
    url = es_url.rstrip("/") + f"/{index}"
    # Create if not exists (PUT is idempotent, but will error if exists with different settings)
    mapping = {
        "mappings": {
            "properties": {
                "text": {"type": "text"},
                "title": {"type": "text"},
                "doc_type": {"type": "keyword"},
                "vendor": {"type": "keyword"},
                "equipment": {"type": "keyword"},
                "site": {"type": "keyword"},
                "building": {"type": "keyword"},
                "version": {"type": "keyword"},
                "effective_date": {"type": "date"},
                "chunk_id": {"type": "keyword"},
                "embedding": {
                    "type": "dense_vector",
                    "dims": dims,
                    "index": True,
                    "similarity": "cosine"
                }
            }
        }
    }
    r = requests.put(url, json=mapping, timeout=30)
    if r.status_code in (200, 201):
        return True
    # If exists already, ignore
    if r.status_code == 400 and "resource_already_exists_exception" in r.text:
        return False
    # If exists, ES may return 400 with already exists depending on version
    if r.status_code == 400 and "resource_already_exists" in r.text:
        return False
    r.raise_for_status()
    return False

def bulk_index(es_url: str, index: str, docs, timeout: int = 60):
    """Use _bulk API. docs is iterable of (id, source_dict)."""
    url = es_url.rstrip("/") + "/_bulk"
    lines = []
    for doc_id, src in docs:
        meta = {"index": {"_index": index}}
        if doc_id:
            meta["index"]["_id"] = doc_id
        lines.append(json.dumps(meta, ensure_ascii=False))
        lines.append(json.dumps(src, ensure_ascii=False))
    body = "\n".join(lines) + "\n"
    r = requests.post(url, data=body.encode("utf-8"), headers={"Content-Type": "application/x-ndjson"}, timeout=timeout)
    r.raise_for_status()
    data = r.json()
    if data.get("errors"):
        # Print first few errors for debugging
        items = data.get("items", [])
        errs = [it["index"].get("error") for it in items if it.get("index", {}).get("error")]
        raise RuntimeError(f"Bulk indexing had errors. Sample: {errs[:3]}")
    return len(docs)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--es-url", default="http://localhost:9200")
    ap.add_argument("--index", default="epms_docs")
    ap.add_argument("--ollama-url", default="http://localhost:11434")
    ap.add_argument("--embed-model", default="nomic-embed-text")
    ap.add_argument("--header-base-row", type=int, default=4)
    ap.add_argument("--header-human-row", type=int, default=5)
    ap.add_argument("--header-suffix-row", type=int, default=7)
    ap.add_argument("--data-start-row", type=int, default=8)
    ap.add_argument("--batch-size", type=int, default=200)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    base = [clean(ws.cell(args.header_base_row, c).value) for c in range(1, max_col + 1)]
    human = [clean(ws.cell(args.header_human_row, c).value) for c in range(1, max_col + 1)]
    suffix = [clean(ws.cell(args.header_suffix_row, c).value) for c in range(1, max_col + 1)]

    colnames = []
    for c in range(1, max_col + 1):
        b, h, s = base[c - 1], human[c - 1], suffix[c - 1]
        if c <= 5:
            colnames.append(b if b else f"col{c}")
        else:
            parts = []
            if b: parts.append(b)
            if h: parts.append(h)
            if s: parts.append(f"({s})")
            colnames.append(" ".join(parts) if parts else f"col{c}")

    # Determine embedding dims from a short test
    test_emb = get_embedding(args.ollama_url, args.embed_model, "차원 확인 테스트")
    dims = len(test_emb)
    print(f"[OK] Embedding dims = {dims}")

    # Ensure index exists with correct dims
    created = ensure_index(args.es_url, args.index, dims)
    if created:
        print(f"[OK] Created index: {args.index}")
    else:
        print(f"[OK] Index exists (or already created): {args.index}")

    def iter_docs():
        for r in range(args.data_start_row, max_row + 1):
            no = ws.cell(r, 1).value
            if no is None:
                continue
            if isinstance(no, str) and not no.strip().isdigit():
                continue

            row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            row = dict(zip(colnames, row_vals))

            no = row.get("NO")
            plc = row.get("PLC No")
            item = row.get("항목")
            panel = row.get("Panel Name")
            base_addr = row.get("기준 Address") or row.get("기준\nAddress")

            signals = []
            for col in colnames[5:]:
                v = row.get(col)
                if not is_blank(v):
                    signals.append(f"- {col}: {v}")

            text = "\n".join([
                "[IO Tag 정의]",
                f"NO: {no}",
                f"PLC: {plc}",
                f"항목(Item): {item}",
                f"Panel: {panel}",
                f"기준 Address: {base_addr}",
                "신호/주소:",
                *signals
            ])

            emb = get_embedding(args.ollama_url, args.embed_model, text)
            src = {
                "chunk_id": f"io_tag_{int(no)}" if not is_blank(no) else None,
                "doc_type": "io_taglist",
                "site": "아산병원",
                "title": f"PLC IO Tag - {item}",
                "text": text,
                "embedding": emb
            }
            doc_id = src["chunk_id"]
            yield doc_id, src

    batch = []
    total = 0
    for doc in iter_docs():
        batch.append(doc)
        if len(batch) >= args.batch_size:
            bulk_index(args.es_url, args.index, batch)
            total += len(batch)
            print(f"[OK] Indexed {total} docs...")
            batch = []
            time.sleep(0.2)

    if batch:
        bulk_index(args.es_url, args.index, batch)
        total += len(batch)

    print(f"[DONE] Total indexed docs: {total}")

if __name__ == "__main__":
    main()
